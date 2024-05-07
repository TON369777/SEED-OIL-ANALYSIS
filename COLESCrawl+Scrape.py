### BEFORE RUNNING, PLEASE PREPARE A SPREADSHEET WITH 2 columns DETAILS OF CATEGORIES OF PRODUCTS TO BE OBTAINED.
### EXAMPLE 1: Pantry pantry
### EXAMPLE 2: Baby baby
### 'Pantry' is for categorising the product data in output csv
### 'pantry' is the category as per COLES website url e.g. https://www.coles.com.au/browse/pantry

# Load Libraries
import datetime
import openpyxl as xl
import time
import csv
import random
from urllib.parse import urljoin

from selenium import webdriver
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

today = datetime.date.today()
start = datetime.datetime.now()
print(start.strftime("%y-%m-%d %H:%M:%S"))

sleep_duration = random.uniform(4,6)

# Initial Run to set up CSV file
header = ['Category','Product Name', 'Ingredients']

with open(f'{today}COLESProductInfo.csv', 'w', newline='', encoding='UTF8') as f:
    writer = csv.writer(f)
    writer.writerow(header)

# Function to set up firefox with ublock on startup
def setup_firefox_with_ublock(xpi_path):
    # Set up Firefox options
    firefox_options = Options()

    # Initialize the WebDriver with the specified options
    driver = webdriver.Firefox(service=Service(), options=firefox_options)

    # Add the uBlock Origin extension after the WebDriver session has started
    driver.install_addon(xpi_path, temporary=True)

    return driver

# Replace 'path_to_ublock.xpi' with the path to your downloaded uBlock Origin .xpi file
ublock_xpi_path = "path_to_ublock.xpi"

# Initialize a set to store visited URLs
visited_urls = set()

# Function to obtain all HREF on current page
def extract_href():
    while True:
        time.sleep(5)
        href_elements = driver.find_elements(By.CSS_SELECTOR, "a.product__link")
        href_list = [element.get_attribute("href") for element in href_elements]
        base_url = 'https://www.coles.com.au/'
        time.sleep(sleep_duration)
        for href in href_list:
            product_page = urljoin(base_url, href)
            print(product_page)

            # GOES TO INDIVIDUAL PRODUCT PAGE AND EXTRACT DESIRED DATA
            if product_page.startswith("https://www.coles.com.au/product/") and product_page not in visited_urls:
                driver.execute_script("window.open('about:blank', '_blank');")
                driver.switch_to.window(driver.window_handles[1])
                driver.get(product_page)
                time.sleep(1)
                product_parse()
                time.sleep(1)
                driver.close()
                # switch back to the original tab and looks for the next one
                driver.switch_to.window(driver.window_handles[0])
                time.sleep(1)
                if product_page not in visited_urls:
                    visited_urls.add(product_page)

                time.sleep(1)

    page_numbers = driver.find_element(By.CSS_SELECTOR, "div[data-testid='pagination-info']").text.split('-')[1].split(" ")
    print(page_numbers)

    if page_numbers[1] == page_numbers[3]: ## loop will break here once the last page of products has been evaluated
        print('No more pages')
        break
    else:
        next_page = driver.find_element(By.CSS_SELECTOR, '#pagination-button-next').click()
        print('Going Next Page')

# Function for Product Page Parsing
def product_parse():
    time.sleep(1)
    try:
        # Wait for ingredient click element
        wait = WebDriverWait(driver, 10)  # Wait up to 10 seconds
        click_element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '.product-details')))
    except TimeoutException:
        print("Timeout occurred while waiting for product detail element")
        pass

    name = driver.find_element(By.CSS_SELECTOR, 'h1.LinesEllipsis')
    name = name.text
    print(name)

    try:
        expand_ingred = driver.find_element(By.CSS_SELECTOR, '#ingredients-label').click()
    except:
        pass

    try:
        ingredient = driver.find_element(By.CSS_SELECTOR, 'div#ingredients-control')
        ingredient = ingredient.text
    except:
        ingredient = ''
    print(ingredient)

    # Write to CSV File
    # header = ['Category', 'Name', 'Ingredients', 'URL']
    data = [category, name, ingredient]

    # Append data to CSV for subsequent runs
    with open(f'{today}COLESProductInfo.csv', 'a+', newline='', encoding='UTF8') as f:
        writer = csv.writer(f)
        writer.writerow(data)
    # time.sleep(1)



##START OPERATION##

# Load Batch Details
wb = xl.load_workbook('COLESbatchdetails.xlsx') ## MUST HAVE PREPARED XLSX file with DETAILS OF product categories ##
ws = wb.active

# Setup Firefox with uBlock
driver = setup_firefox_with_ublock(ublock_xpi_path)
print("ublock set up")
# Now you can navigate to any page and uBlock will be active
driver.get('https://www.coles.com.au/')
time.sleep(5)

# Running through batch details
for row in range (1, ws.max_row):
    #OPTION A:  Will crawl through product categories according to input in spreadsheet#####################################
    category = ws.cell(row = row, column = 1).value
    class_ = ws.cell(row = row, column = 2).value
    url = f'https://www.coles.com.au/browse/{class_}'
    ########################################################################################################################

    #OPTION B:  IF PROGRAM IS INTERRUPTED, CAN CONTINUE WITH INPUTTING DETAILS HERE AND HASHING THE PREVIOUS SECTION##############
    # category = 'Baby'
    # class_ = 'baby'
    # url = 'https://www.coles.com.au/browse/baby'  ## CHANGE THE FILE NAME BEFORE RE RUNNING FILE ##
    ######################################################################################################################

    print("Extracting product info for:", category)
    driver.get(url)
    time.sleep(5)
    extract_href()
    time.sleep(5)

print("CRAWL AND SCRAPE OPERATION COMPLETE")
finish = datetime.datetime.now()
print(finish.strftime("%y-%m-%d %H:%M:%S"))
# close WebDriver
driver.quit()